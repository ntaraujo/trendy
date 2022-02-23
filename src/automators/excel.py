from gooey import local_resource_path

if __name__ == '__main__':
    from sys import path as sys_path

    sys_path.insert(0, local_resource_path(""))

from utils import msg, global_path, scheduled, open
from openpyxl.styles.colors import Color as XlsColor
from openpyxl.styles.fills import PatternFill as XlsPatternFill
from openpyxl.styles import Font as XlsFont
from openpyxl.styles.borders import Border as XlsBorder
from openpyxl.styles.borders import Side as XlsSide
from openpyxl.styles.borders import BORDER_THIN as XLS_BORDER_THIN


class Excel:
    def __init__(self):
        self.macros_file = None
        self.file = None
        self.path = None
        self.app = None
    
    @scheduled
    def open_app(self):
        import xlwings as xw

        if xw.apps.count < 2:
            msg("Iniciando uma instância do Excel para uso do programa")
            new_app = xw.App()
            new_app.activate()
            # new_app.visible = True
            self.app = new_app
        else:
            pids = sorted(xw.apps.keys())
            self.app = xw.apps[pids[-1]]

    @scheduled
    def open_macros(self):
        msg("Abrindo macros")

        macros_path = global_path("resources/macros.xlsb")
        self.macros_file = self.app.books.open(macros_path)
        self.macros_file.activate()

    @scheduled
    def open(self, path=None):
        if path is None:
            msg("Abrindo nova pasta do Excel")
            self.file = self.app.books.add()
        else:
            from os.path import abspath
            path = abspath(path)
            msg(f'Abrindo pasta do Excel em "{path}"')
            self.file = self.app.books.open(path)

        self.file.activate()
        # self.file.app.visible = not compiled

        self.path = path

    @scheduled
    def close(self):
        if self.path is not None:
            msg("Salvando e fechando o arquivo")

            self.file.save()
        else:
            msg("Fechando o arquivo")
        self.file.close()

    @scheduled
    def run(self, macro, *args, **kwargs):
        msg(f'Executando a macro "{macro}"')

        self.macros_file.macro(macro)(*args, **kwargs)

    @scheduled
    def insert(self, cells):
        msg("Inserindo linhas")

        used = self.file.sheets.active.used_range
        if used.value is None:
            used.value = cells
        else:
            self.file.sheets.active.range((used.last_cell.row + 1, 1)).value = cells

    @staticmethod
    def merge_across(cells):
        msg("Células mescladas")

        cells.merge(across=True)

    @staticmethod
    def bold(cells):
        msg("Células em negrito")

        cells.font.bold = True

    @scheduled
    def center(self, cells):
        msg("Células centralizadas")

        cells.select()
        self.run("center_selected")

    @staticmethod
    def font_color(cells, r, g, b):
        msg("Células com fonte colorida")

        cells.font.color = (r, g, b)

    @staticmethod
    def color(cells, r, g, b):
        msg("Células coloridas")

        cells.color = (r, g, b)

    @scheduled
    def new_sheet(self, name=None):
        msg("Adicionando nova planilha")

        if name is not None and len(name) > 31:
            name = name[:31]

        self.file.sheets.add(name, after=self.file.sheets.active)

    @scheduled
    def save(self, file_path=None):
        msg("Salvando arquivo")

        self.file.save(file_path)

    @staticmethod
    def xls_file_vertical_search(value, file_or_path, lookup_col, *return_cols):
        msg(f'Procurando "{value}"')

        if type(file_or_path) == str:
            from openpyxl import load_workbook

            ws = load_workbook(file_or_path).active
        else:
            ws = file_or_path
        for row in ws.values:
            if row is None:
                return
            if value in (row[lookup_col - 1] or ''):
                yield [row[return_col - 1] for return_col in return_cols]
    
    @staticmethod
    def xls_workbook(file_path=None):
        msg(f'Abrindo "{file_path}"')
        
        if file_path:
            from openpyxl import load_workbook

            return load_workbook(file_path)
        else:
            from openpyxl import Workbook

            return Workbook()
    
    @staticmethod
    def xls_color(cell, rgb):
        msg("Célula colorida")

        cell.fill = XlsPatternFill(patternType='solid', fgColor=XlsColor(rgb=rgb))
    
    @staticmethod
    def xls_bold(cell):
        msg("Célula em negrito")

        cell.font = XlsFont(bold=True)
    
    @staticmethod
    def xls_thin_border(cell):
        msg("Célula com todas as bordas finas")
        cell.border = XlsBorder(
            left=XlsSide(border_style=XLS_BORDER_THIN, color='00000000'),
            right=XlsSide(border_style=XLS_BORDER_THIN, color='00000000'),
            top=XlsSide(border_style=XLS_BORDER_THIN, color='00000000'),
            bottom=XlsSide(border_style=XLS_BORDER_THIN, color='00000000')
        )
    
    @staticmethod
    def xls_on_back_range(sheet, rows, columns, *cell_actions):
        msg("Ações nas últimas linhas")

        max_row = sheet.max_row
        
        for cell_action in cell_actions:
            if type(cell_action) == tuple:
                args = cell_action[1]
                if len(cell_action) == 3:
                    kwargs = cell_action[2]
                else:
                    kwargs = {}
                cell_action = cell_action[0]
            else:
                args = []
                kwargs = {}
            
            for r in range(rows):
                for c in range(columns):
                    cell_action(sheet.cell(max_row-r, c+1), *args, **kwargs)

    @staticmethod
    def csv_file_vertical_search(value, file_or_path, lookup_col, *return_cols):
        msg(f'Procurando "{value}"')

        if type(file_or_path) == str:
            from csv import reader

            file = reader(open(file_or_path))
        else:
            file = file_or_path
        for row in file:
            if value in (row[lookup_col - 1] or ''):
                yield [row[return_col - 1] for return_col in return_cols]
    
    @staticmethod
    def get_csv_reader(file_path, open_kwargs=None, reader_kwargs=None):
        msg(f'Abrindo "{file_path}"')

        from csv import reader

        open_kwargs = open_kwargs or {}
        reader_kwargs = reader_kwargs or {}
        return reader(open(file_path, **open_kwargs), **reader_kwargs)
    
    @scheduled
    def delete_sheet(self, identifier):
        msg(f'Deletando planilha {repr(identifier)}')

        self.file.sheets[identifier].delete()
    
    @scheduled
    def on_back_range(self, rows, columns, *cell_actions):
        msg("Ações nas últimas linhas")

        used = self.file.sheets.active.used_range
        bottom_left = self.file.sheets.active.range((used.last_cell.row, 1))
        top_right = bottom_left.offset(-rows + 1, columns - 1)
        back_range = self.file.sheets.active.range(bottom_left.address + ":" + top_right.address)
        
        for cell_action in cell_actions:
            if type(cell_action) == tuple:
                args = cell_action[1]
                if len(cell_action) == 3:
                    kwargs = cell_action[2]
                else:
                    kwargs = {}
                cell_action[0](back_range, *args, **kwargs)
            else:
                cell_action(back_range)
    
    @scheduled
    def assign(self, cell_ref, value):
        msg("Atribuindo valor à célula")
        
        self.file.sheets.active.range(cell_ref).value = value


if __name__ == "__main__":
    excel = Excel()
    excel.open()
    excel.insert([['Pedido', 'Status', 'Tipo Ped.', 'Est', 'Dt Refer', 'Prev Emb', 'Pré-Data', 'NF', 'Dt Saída',
                   'Modelo', 'Descrição', 'Qt Pares', 'Vl Líq', 'Nr Ordem', 'CR', 'CO', 'EF'],
                  ['TOTAL', '\xa0', '\xa0', '\xa0', '\xa0', '\xa0', '\xa0', '\xa0', '\xa0', '\xa0', '\xa0', '768',
                   '52.537,08', '\xa0', '\xa0', '\xa0', '\xa0'],
                  ['490766801', 'Faturado', 'Normal', '20', '26/01/2022', '03/01/2022', '?', '5226220', '29/01/2022',
                   '33423', 'MELISSA PAPETE WIDE AD', '24', '2.451,84', '\xa0', 'L', 'L', 'L'],
                  ['490766802', 'Faturado', 'Normal', '21', '19/01/2022', '03/01/2022', '?', '1529257', '19/01/2022',
                   '33427', 'MELISSA SHINY HEEL AD', '24', '2.838,96', '\xa0', 'L', 'L', 'L'],
                  ['490766803', 'Faturado', 'Normal', '20', '19/01/2022', '03/01/2022', '?', '5222890', '19/01/2022',
                   '33429', 'MELISSA SHINY AD', '36', '1.935,36', '\xa0', 'L', 'L', 'L'],
                  ['490766804', 'Faturado', 'Normal', '20', '19/01/2022', '03/01/2022', '?', '5222889', '19/01/2022',
                   '33431', 'MELISSA BRIGHTNESS AD', '36', '3.290,40', '\xa0', 'L', 'L', 'L'],
                  ['490766805', 'Faturado', 'Normal', '21', '17/01/2022', '03/01/2022', '?', '1528851', '18/01/2022',
                   '33521', 'MINI MELISSA POSSESSION SHINY INF', '24', '1.935,60', '\xa0', 'L', 'L', 'L'],
                  ['490766806', 'Faturado', 'Normal', '21', '17/01/2022', '03/01/2022', '?', '1528541', '18/01/2022',
                   '33522', 'MINI MELISSA POSSESSION SHINY BB', '24', '1.419,36', '\xa0', 'L', 'L', 'L'],
                  ['490766807', 'Faturado', 'Normal', '20', '17/01/2022', '03/01/2022', '?', '5220779', '18/01/2022',
                   '33528', 'MELISSA SUN LONG BEACH AD', '36', '1.161,00', '\xa0', 'L', 'L', 'L'],
                  ['490766808', 'Faturado', 'Normal', '20', '19/01/2022', '03/01/2022', '?', '5222986', '20/01/2022',
                   '33530', 'MELISSA SUN RODEO AD', '30', '1.451,70', '\xa0', 'L', 'L', 'L'],
                  ['490766809', 'Programado', 'Normal', '40', '26/02/2022', '03/01/2022', '?', '\xa0', '?', '33531',
                   'MELISSA FLIP FLOP FREE AD', '24', '2.193,60', '\xa0', 'L', 'L', 'L'],
                  ['490766810', 'Faturado', 'Normal', '21', '17/01/2022', '03/01/2022', '?', '1528808', '18/01/2022',
                   '33538', 'MELISSA SOLAR II + BOBO AD', '12', '1.225,92', '\xa0', 'L', 'L', 'L'],
                  ['490766811', 'Faturado', 'Normal', '21', '17/01/2022', '03/01/2022', '?', '1528705', '18/01/2022',
                   '33539', 'MELISSA HARMONIC CHROME IX AD', '36', '2.129,04', '\xa0', 'L', 'L', 'L'],
                  ['490766812', 'Faturado', 'Normal', '21', '24/01/2022', '03/01/2022', '?', '1532188', '26/01/2022',
                   '33542', 'MELISSA MULE III AD', '12', '1.612,92', '\xa0', 'L', 'L', 'L'],
                  ['490766813', 'Faturado', 'Normal', '21', '17/01/2022', '03/01/2022', '?', '1528832', '18/01/2022',
                   '33546', 'MINI MELISSA MAR SANDAL JELLY POP INF', '24', '1.935,60', '\xa0', 'L', 'L', 'L'],
                  ['490766814', 'Faturado', 'Normal', '20', '24/01/2022', '03/01/2022', '?', '5225125', '25/01/2022',
                   '33547', 'MELISSA BIKINI STRIPE AD', '24', '1.290,48', '\xa0', 'L', 'L', 'L'],
                  ['490766815', 'Faturado', 'Normal', '20', '26/01/2022', '03/01/2022', '?', '5226223', '29/01/2022',
                   '33557', 'MELISSA SUN CITY WALK AD', '18', '871,02', '\xa0', 'L', 'L', 'L'],
                  ['490766816', 'Faturado', 'Normal', '21', '17/01/2022', '03/01/2022', '?', '1528680', '18/01/2022',
                   '33559', 'MINI MELISSA DORA III BB', '18', '967,86', '\xa0', 'L', 'L', 'L'],
                  ['490766817', 'Faturado', 'Normal', '21', '21/01/2022', '03/01/2022', '21/01/2022', '1530912',
                   '21/01/2022', '33571', 'MELISSA THE REAL JELLY SANDAL AD', '12', '838,80', '\xa0', 'L', 'L', 'L'],
                  ['490766818', 'Faturado', 'Normal', '21', '17/01/2022', '03/01/2022', '?', '1528793', '18/01/2022',
                   '33580', 'MINI MELISSA SUNNY BB', '30', '1.613,10', '\xa0', 'L', 'L', 'L'],
                  ['490766819', 'Faturado', 'Normal', '21', '17/01/2022', '03/01/2022', '?', '1528611', '18/01/2022',
                   '33587', 'MELISSA FUNKY AD', '24', '3.484,08', '\xa0', 'L', 'L', 'L'],
                  ['490766820', 'Faturado', 'Normal', '20', '24/01/2022', '03/01/2022', '?', '5225139', '25/01/2022',
                   '33614', 'MELISSA FLIP FLOP SLIM III AD', '18', '1.258,02', '\xa0', 'L', 'L', 'L'],
                  ['490766821', 'Faturado', 'Normal', '20', '24/01/2022', '03/01/2022', '?', '5225124', '25/01/2022',
                   '33617', 'MINI MELISSA COSMIC SANDAL INF', '30', '2.580,90', '\xa0', 'L', 'L', 'L'],
                  ['490766822', 'Faturado', 'Normal', '21', '24/01/2022', '03/01/2022', '?', '1532239', '24/01/2022',
                   '33634', 'MELISSA SEDUCTION VI AD', '12', '967,80', '\xa0', 'L', 'L', 'L'],
                  ['490766823', 'Faturado', 'Normal', '21', '21/01/2022', '03/01/2022', '21/01/2022', '1530728',
                   '21/01/2022', '33646', 'MELISSA THE REAL JELLY SLIDE AD', '12', '774,24', '\xa0', 'L', 'L', 'L'],
                  ['490766824', 'Faturado', 'Normal', '20', '24/01/2022', '03/01/2022', '?', '5225077', '24/01/2022',
                   '33656', 'MELISSA DARE STRAP + CAMILA COUTINHO AD', '30', '2.903,10', '\xa0', 'L', 'L', 'L'],
                  ['490766825', 'Faturado', 'Normal', '21', '17/01/2022', '03/01/2022', '?', '1528769', '18/01/2022',
                   '33657', 'MELISSA T-BAR STRAP + CAMILA COUTINHO AD', '24', '1.548,48', '\xa0', 'L', 'L', 'L'],
                  ['490766826', 'Faturado', 'Normal', '21', '17/01/2022', '03/01/2022', '?', '1528575', '18/01/2022',
                   '33682', 'MINI MELISSA ULTRAGIRL SWEET X BB', '24', '1.548,48', '\xa0', 'L', 'L', 'L'],
                  ['490766827', 'Faturado', 'Normal', '20', '17/01/2022', '03/01/2022', '?', '5220778', '18/01/2022',
                   '33694', 'MELISSA SUN VENICE SHINY AD', '36', '1.355,04', '\xa0', 'L', 'L', 'L'],
                  ['490766828', 'Faturado', 'Normal', '21   ', '24/01/2022', '03/01/2022', '?', '1532341', '26/01/2022',
                   '33771', 'MELISSA AIRBUBBLE FLIP FLOP AD', '24', '1.677,60', '\xa0', 'L', 'L', 'L'],
                  ['490766829', 'Programado', 'Normal', '40', '26/02/2022', '03/01/2022', '?', '\xa0', '?', '33772',
                   'MELISSA FREE PLATFORM AD', '18', '2.129,22', '\xa0', 'L', 'L', 'L'],
                  ['490767001', 'Faturado', 'Normal', '20', '28/12/2021', '03/01/2022', '?', '5213549', '30/12/2021',
                   '34102', 'MINIATURA MELISSA CORACAO XIII SP', '60', '561,00', '\xa0', 'L', 'L', 'L'],
                  ['490767002', 'Faturado', 'Normal', '20', '24/01/2022', '03/01/2022', '?', '5225144', '25/01/2022',
                   '34305', 'MELISSA SUN SANTA MONICA II', '12', '586,56', '\xa0', 'L', 'L', 'L']])
    excel.insert([['Pedido', 'Status', 'Tipo Ped.', 'Est', 'Dt Refer', 'Prev Emb', 'Pré-Data', 'NF', 'Dt Saída',
                   'Modelo', 'Descrição', 'Qt Pares', 'Vl Líq', 'Nr Ordem', 'CR', 'CO', 'EF'],
                  ['TOTAL', '\xa0', '\xa0', '\xa0', '\xa0', '\xa0', '\xa0', '\xa0', '\xa0', '\xa0', '\xa0', '768',
                   '52.537,08', '\xa0', '\xa0', '\xa0', '\xa0'],
                  ['490766801', 'Faturado', 'Normal', '20', '26/01/2022', '03/01/2022', '?', '5226220', '29/01/2022',
                   '33423', 'MELISSA PAPETE WIDE AD', '24', '2.451,84', '\xa0', 'L', 'L', 'L'],
                  ['490766802', 'Faturado', 'Normal', '21', '19/01/2022', '03/01/2022', '?', '1529257', '19/01/2022',
                   '33427', 'MELISSA SHINY HEEL AD', '24', '2.838,96', '\xa0', 'L', 'L', 'L'],
                  ['490766803', 'Faturado', 'Normal', '20', '19/01/2022', '03/01/2022', '?', '5222890', '19/01/2022',
                   '33429', 'MELISSA SHINY AD', '36', '1.935,36', '\xa0', 'L', 'L', 'L'],
                  ['490766804', 'Faturado', 'Normal', '20', '19/01/2022', '03/01/2022', '?', '5222889', '19/01/2022',
                   '33431', 'MELISSA BRIGHTNESS AD', '36', '3.290,40', '\xa0', 'L', 'L', 'L'],
                  ['490766805', 'Faturado', 'Normal', '21', '17/01/2022', '03/01/2022', '?', '1528851', '18/01/2022',
                   '33521', 'MINI MELISSA POSSESSION SHINY INF', '24', '1.935,60', '\xa0', 'L', 'L', 'L'],
                  ['490766806', 'Faturado', 'Normal', '21', '17/01/2022', '03/01/2022', '?', '1528541', '18/01/2022',
                   '33522', 'MINI MELISSA POSSESSION SHINY BB', '24', '1.419,36', '\xa0', 'L', 'L', 'L'],
                  ['490766807', 'Faturado', 'Normal', '20', '17/01/2022', '03/01/2022', '?', '5220779', '18/01/2022',
                   '33528', 'MELISSA SUN LONG BEACH AD', '36', '1.161,00', '\xa0', 'L', 'L', 'L'],
                  ['490766808', 'Faturado', 'Normal', '20', '19/01/2022', '03/01/2022', '?', '5222986', '20/01/2022',
                   '33530', 'MELISSA SUN RODEO AD', '30', '1.451,70', '\xa0', 'L', 'L', 'L'],
                  ['490766809', 'Programado', 'Normal', '40', '26/02/2022', '03/01/2022', '?', '\xa0', '?', '33531',
                   'MELISSA FLIP FLOP FREE AD', '24', '2.193,60', '\xa0', 'L', 'L', 'L'],
                  ['490766810', 'Faturado', 'Normal', '21', '17/01/2022', '03/01/2022', '?', '1528808', '18/01/2022',
                   '33538', 'MELISSA SOLAR II + BOBO AD', '12', '1.225,92', '\xa0', 'L', 'L', 'L'],
                  ['490766811', 'Faturado', 'Normal', '21', '17/01/2022', '03/01/2022', '?', '1528705', '18/01/2022',
                   '33539', 'MELISSA HARMONIC CHROME IX AD', '36', '2.129,04', '\xa0', 'L', 'L', 'L'],
                  ['490766812', 'Faturado', 'Normal', '21', '24/01/2022', '03/01/2022', '?', '1532188', '26/01/2022',
                   '33542', 'MELISSA MULE III AD', '12', '1.612,92', '\xa0', 'L', 'L', 'L'],
                  ['490766813', 'Faturado', 'Normal', '21', '17/01/2022', '03/01/2022', '?', '1528832', '18/01/2022',
                   '33546', 'MINI MELISSA MAR SANDAL JELLY POP INF', '24', '1.935,60', '\xa0', 'L', 'L', 'L'],
                  ['490766814', 'Faturado', 'Normal', '20', '24/01/2022', '03/01/2022', '?', '5225125', '25/01/2022',
                   '33547', 'MELISSA BIKINI STRIPE AD', '24', '1.290,48', '\xa0', 'L', 'L', 'L'],
                  ['490766815', 'Faturado', 'Normal', '20', '26/01/2022', '03/01/2022', '?', '5226223', '29/01/2022',
                   '33557', 'MELISSA SUN CITY WALK AD', '18', '871,02', '\xa0', 'L', 'L', 'L'],
                  ['490766816', 'Faturado', 'Normal', '21', '17/01/2022', '03/01/2022', '?', '1528680', '18/01/2022',
                   '33559', 'MINI MELISSA DORA III BB', '18', '967,86', '\xa0', 'L', 'L', 'L'],
                  ['490766817', 'Faturado', 'Normal', '21', '21/01/2022', '03/01/2022', '21/01/2022', '1530912',
                   '21/01/2022', '33571', 'MELISSA THE REAL JELLY SANDAL AD', '12', '838,80', '\xa0', 'L', 'L', 'L'],
                  ['490766818', 'Faturado', 'Normal', '21', '17/01/2022', '03/01/2022', '?', '1528793', '18/01/2022',
                   '33580', 'MINI MELISSA SUNNY BB', '30', '1.613,10', '\xa0', 'L', 'L', 'L'],
                  ['490766819', 'Faturado', 'Normal', '21', '17/01/2022', '03/01/2022', '?', '1528611', '18/01/2022',
                   '33587', 'MELISSA FUNKY AD', '24', '3.484,08', '\xa0', 'L', 'L', 'L'],
                  ['490766820', 'Faturado', 'Normal', '20', '24/01/2022', '03/01/2022', '?', '5225139', '25/01/2022',
                   '33614', 'MELISSA FLIP FLOP SLIM III AD', '18', '1.258,02', '\xa0', 'L', 'L', 'L'],
                  ['490766821', 'Faturado', 'Normal', '20', '24/01/2022', '03/01/2022', '?', '5225124', '25/01/2022',
                   '33617', 'MINI MELISSA COSMIC SANDAL INF', '30', '2.580,90', '\xa0', 'L', 'L', 'L'],
                  ['490766822', 'Faturado', 'Normal', '21', '24/01/2022', '03/01/2022', '?', '1532239', '24/01/2022',
                   '33634', 'MELISSA SEDUCTION VI AD', '12', '967,80', '\xa0', 'L', 'L', 'L'],
                  ['490766823', 'Faturado', 'Normal', '21', '21/01/2022', '03/01/2022', '21/01/2022', '1530728',
                   '21/01/2022', '33646', 'MELISSA THE REAL JELLY SLIDE AD', '12', '774,24', '\xa0', 'L', 'L', 'L'],
                  ['490766824', 'Faturado', 'Normal', '20', '24/01/2022', '03/01/2022', '?', '5225077', '24/01/2022',
                   '33656', 'MELISSA DARE STRAP + CAMILA COUTINHO AD', '30', '2.903,10', '\xa0', 'L', 'L', 'L'],
                  ['490766825', 'Faturado', 'Normal', '21', '17/01/2022', '03/01/2022', '?', '1528769', '18/01/2022',
                   '33657', 'MELISSA T-BAR STRAP + CAMILA COUTINHO AD', '24', '1.548,48', '\xa0', 'L', 'L', 'L'],
                  ['490766826', 'Faturado', 'Normal', '21', '17/01/2022', '03/01/2022', '?', '1528575', '18/01/2022',
                   '33682', 'MINI MELISSA ULTRAGIRL SWEET X BB', '24', '1.548,48', '\xa0', 'L', 'L', 'L'],
                  ['490766827', 'Faturado', 'Normal', '20', '17/01/2022', '03/01/2022', '?', '5220778', '18/01/2022',
                   '33694', 'MELISSA SUN VENICE SHINY AD', '36', '1.355,04', '\xa0', 'L', 'L', 'L'],
                  ['490766828', 'Faturado', 'Normal', '21   ', '24/01/2022', '03/01/2022', '?', '1532341', '26/01/2022',
                   '33771', 'MELISSA AIRBUBBLE FLIP FLOP AD', '24', '1.677,60', '\xa0', 'L', 'L', 'L'],
                  ['490766829', 'Programado', 'Normal', '40', '26/02/2022', '03/01/2022', '?', '\xa0', '?', '33772',
                   'MELISSA FREE PLATFORM AD', '18', '2.129,22', '\xa0', 'L', 'L', 'L'],
                  ['490767001', 'Faturado', 'Normal', '20', '28/12/2021', '03/01/2022', '?', '5213549', '30/12/2021',
                   '34102', 'MINIATURA MELISSA CORACAO XIII SP', '60', '561,00', '\xa0', 'L', 'L', 'L'],
                  ['490767002', 'Faturado', 'Normal', '20', '24/01/2022', '03/01/2022', '?', '5225144', '25/01/2022',
                   '34305', 'MELISSA SUN SANTA MONICA II', '12', '586,56', '\xa0', 'L', 'L', 'L']])
    # excel.close()
